import openpyxl, json
from datetime import datetime

wb = openpyxl.load_workbook("brand_origin_lookup.xlsx", data_only=True)
ws = wb["Sheet1"]

origin = {}
private_label = []
spinneys_supplement = []

for r in range(2, ws.max_row + 1):

    brand = ws.cell(row=r, column=2).value
    if isinstance(brand, datetime):
        print(f"Skipping row {r}: invalid brand {brand!r}")
        continue
    brand = str(brand).strip()

    am = ws.cell(row=r, column=3).value or 0
    pm = ws.cell(row=r, column=4).value or 0
    sp = ws.cell(row=r, column=5).value or 0
    country = ws.cell(row=r, column=8).value
    conf = ws.cell(row=r, column=9).value
    if not brand:
        continue
    if conf == "N/A":
        private_label.append(brand)
    elif country:
        origin[brand] = country
    if am == 0 and pm == 0 and sp > 0 and conf != "N/A":
        spinneys_supplement.append(brand)

json.dump(origin, open("../lib/etl/brandOrigin.json", "w"), ensure_ascii=False, indent=2)
json.dump(private_label, open("../lib/etl/privateLabelBrands.json", "w"), ensure_ascii=False, indent=2)
json.dump(sorted(spinneys_supplement), open("../lib/etl/spinneysSupplementBrands.json", "w"), ensure_ascii=False, indent=2)
